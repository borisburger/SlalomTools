import os
import json
import base64
import time
import asyncio
import platform
from datetime import datetime
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, UploadFile, File, BackgroundTasks, Query, Request, Response
from fastapi.responses import JSONResponse, RedirectResponse, FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
import requests
import msal
# Import the Google Sheets module
from google_sheets import initiate_auth_flow, complete_auth_flow, get_credentials, fetch_spreadsheet_data, parse_registration_data
# Import rankings module
from rankings import fetch_rankings, get_latest_rankings_folder, format_date_for_folder, get_discipline_file_path, get_download_progress, fetch_skater_database, get_skater_db_progress
import csv
from bs4 import BeautifulSoup

# Load configuration
with open("config.json") as f:
    config = json.load(f)

# Load secrets
with open("secrets.json") as f:
    secrets = json.load(f)

# Paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, config["paths"]["static_dir"])
DEFAULT_DATA_FILE = os.path.join(BASE_DIR, config["paths"]["default_data_file"])
# Update default background path to point to frontend
DEFAULT_BACKGROUND_NAME = config["paths"]["default_background_image"]
FRONTEND_PUBLIC_DIR = os.path.join(os.path.dirname(BASE_DIR), "frontend", "public")
BACKGROUND_IMAGE_PATH = os.path.join(FRONTEND_PUBLIC_DIR, "backgrounds", DEFAULT_BACKGROUND_NAME)
TOKEN_CACHE_FILE = os.path.join(BASE_DIR, "token_cache.json")

# Ensure static directory exists
os.makedirs(STATIC_DIR, exist_ok=True)

# Microsoft Graph settings
AUTHORITY = "https://login.microsoftonline.com/common"
SCOPES = ["Files.Read", "Files.Read.All"]

# FastAPI app
app = FastAPI()

# Handle CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Skater Database API endpoints - moved to top to ensure proper routing
@app.get("/api/skater-db/info")
async def api_skater_db_info():
    """
    API endpoint to get information about the skater database
    """
    db_path = "rankings/skater-db.json"
    print(f"[DEBUG] Checking for skater database at: {db_path}")
    
    if not os.path.exists(db_path):
        print(f"[DEBUG] Skater database file not found at: {db_path}")
        return {
            "exists": False,
            "count": 0,
            "last_updated": None
        }
    
    try:
        # Get file modification time
        last_updated = datetime.fromtimestamp(os.path.getmtime(db_path)).isoformat()
        print(f"[DEBUG] Skater database last modified: {last_updated}")
        
        # Read the skater count from the file
        with open(db_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            print(f"[DEBUG] Successfully read skater database file")
            
            # Get count from total_skaters or length of skaters array
            count = data.get("total_skaters", 0)
            if count == 0 and "skaters" in data:
                count = len(data["skaters"])
            
            print(f"[DEBUG] Skater count: {count}")
            
            return {
                "exists": True,
                "count": count,
                "last_updated": last_updated
            }
            
    except Exception as e:
        print(f"[DEBUG] Error reading skater database: {str(e)}")
        return {
            "exists": True,
            "count": 0,
            "last_updated": None,
            "error": str(e)
        }

@app.post("/api/skater-db/update")
async def api_update_skater_db(background_tasks: BackgroundTasks):
    """
    API endpoint to start skater database download in the background
    """
    try:
        background_tasks.add_task(fetch_skater_database)
        return {"status": "success", "message": "Skater database download started"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/skater-db/progress")
async def api_skater_db_progress():
    """
    API endpoint to get the skater database download progress
    """
    progress = get_skater_db_progress()
    print(f"[DEBUG] Skater DB progress: {progress}")
    return progress

@app.get("/api/skater-db/download", response_class=FileResponse)
async def api_download_skater_db():
    """
    API endpoint to download the skater database JSON file
    """
    db_path = "rankings/skater-db.json"
    if not os.path.exists(db_path):
        return JSONResponse(
            status_code=404,
            content={"status": "error", "message": "Skater database not available. Please download it first."}
        )
    
    return FileResponse(
        path=db_path,
        filename="skater-db.json",
        media_type="application/json"
    )

@app.get("/api/skater-db/info-test")
async def api_skater_db_info_test():
    """
    A test endpoint for the skater database info
    """
    db_path = "rankings/skater-db.json"
    file_exists = os.path.exists(db_path)
    
    result = {
        "exists": file_exists,
        "path": os.path.abspath(db_path) if file_exists else None,
        "file_size": os.path.getsize(db_path) if file_exists else 0
    }
    
    if file_exists:
        try:
            with open(db_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                result["total_skaters"] = data.get("total_skaters", 0)
                result["skaters_array_length"] = len(data.get("skaters", []))
        except Exception as e:
            result["error"] = str(e)
    
    return result

@app.get("/api/skater-db/data")
async def api_skater_db_data():
    """
    API endpoint to get the full skater database as JSON
    This provides the entire database to the frontend for client-side filtering and searching
    """
    db_path = "rankings/skater-db.json"
    print(f"[DEBUG] Loading full skater database from: {db_path}")
    
    if not os.path.exists(db_path):
        print(f"[DEBUG] Skater database file not found at: {db_path}")
        return JSONResponse(
            status_code=404,
            content={"error": "Skater database not found. Please download it first."}
        )
    
    try:
        # Read the full database file
        with open(db_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            print(f"[DEBUG] Successfully loaded full skater database with {len(data.get('skaters', []))} skaters")
            
            # Return the entire data structure
            return data
            
    except Exception as e:
        print(f"[DEBUG] Error reading skater database: {str(e)}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to read skater database: {str(e)}"}
        )

# Determine frontend directory paths
FRONTEND_DIR = os.environ.get("FRONTEND_DIR", None)

# If FRONTEND_DIR is not set, try to detect it
if not FRONTEND_DIR:
    # Check for common frontend build locations relative to the backend directory
    potential_dirs = [
        os.path.join(os.path.dirname(__file__), "..", "frontend", "build"),
        os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"),
        os.path.join(os.path.dirname(__file__), "frontend", "build"),
        os.path.join(os.path.dirname(__file__), "frontend", "dist")
    ]
    
    for dir_path in potential_dirs:
        if os.path.exists(dir_path) and os.path.exists(os.path.join(dir_path, "index.html")):
            FRONTEND_DIR = dir_path
            print(f"Detected frontend directory: {FRONTEND_DIR}")
            break

if FRONTEND_DIR:
    # Mount the frontend static files directory if it exists
    print(f"Frontend directory set to: {FRONTEND_DIR}")
    STATIC_DIR = os.path.join(FRONTEND_DIR, "static")
    if os.path.exists(STATIC_DIR):
        print(f"Static files directory found at: {STATIC_DIR}")
    else:
        print(f"Static files directory not found at: {STATIC_DIR}")
else:
    print("WARNING: Frontend build not found. The React app will not be served.")
    print("Please run the build_frontend script to create the build files.")
    print(f"Checked potential directories:")
    potential_dirs_to_check = [
        os.path.join(os.path.dirname(__file__), "..", "frontend", "build"),
        os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"),
        os.path.join(os.path.dirname(__file__), "frontend", "build"),
        os.path.join(os.path.dirname(__file__), "frontend", "dist")
    ]
    for path in potential_dirs_to_check:
        print(f"  - {path} (exists: {os.path.exists(path)})")

# Application state
state = {
    "current_file": None,    # share URL or local path
    "drive_id": None,        # cached driveId for Graph
    "item_id": None,         # cached itemId for Graph
    "live": {"category": None, "discipline": None, "competitors": [], "category_complete": False},
    "public": {"category": None, "discipline": None, "competitors": [], "category_complete": False, "message": "", "display_mode": "results"},
    "background_url": None,
    "last_file_check": 0,    # timestamp of last file check
    "last_modified": None,   # last modified timestamp of the file
    "auto_refresh_enabled": True, # enable/disable auto refresh
    "auto_refresh_interval": 5   # seconds between checks
}
if os.path.exists(BACKGROUND_IMAGE_PATH):
    # Use the frontend path format for the background URL
    state["background_url"] = f"/backgrounds/{DEFAULT_BACKGROUND_NAME}"

operator_connections = []
public_connections = []

# Authentication state
auth_state = {"flow": None, "token": None, "expires_at": 0, "is_authenticated": False}

# Background tasks
background_tasks = set()

# Registration state
reg_state = {
    "current_sheet_url": None,
    "disciplines": [],
    "skaters": []
}

def load_token_cache():
    """Load the token cache from base64-encoded JSON file if it exists."""
    print(f"Attempting to load token cache from: {TOKEN_CACHE_FILE}")
    if not os.path.exists(TOKEN_CACHE_FILE):
        print("Token cache file does not exist")
        return None
    try:
        encoded = open(TOKEN_CACHE_FILE, 'r').read()
        decoded = base64.b64decode(encoded).decode('utf-8')
        cache = msal.SerializableTokenCache()
        cache.deserialize(decoded)
        print("Successfully loaded token cache")
        return cache
    except Exception as e:
        print(f"Failed to load token cache: {e}")
        return None


def save_token_cache(cache):
    """Save the token cache to base64-encoded JSON file only if we have a valid token."""
    try:
        # Check if we have any accounts in the cache
        accounts = cache.find(msal.TokenCache.CredentialType.ACCOUNT)
        if not accounts:
            print("No accounts in cache, skipping save")
            return

        print(f"Attempting to save token cache to: {TOKEN_CACHE_FILE}")
        serialized = cache.serialize()
        encoded = base64.b64encode(serialized.encode('utf-8')).decode('utf-8')
        with open(TOKEN_CACHE_FILE, 'w') as f:
            f.write(encoded)
        print("Successfully saved token cache")
    except Exception as e:
        print(f"Failed to save token cache: {e}")


def get_user_token():
    """Get an access token using device code flow."""
    global auth_state
    now = time.time()
    if auth_state["token"] and now < auth_state["expires_at"]:
        print("Using cached token from memory")
        return auth_state["token"]

    print("No valid token in memory, attempting to load from cache")
    cache = load_token_cache() or msal.SerializableTokenCache()
    client = msal.PublicClientApplication(
        secrets["microsoft"]["client_id"],
        authority=AUTHORITY,
        token_cache=cache
    )

    accounts = client.get_accounts()
    if accounts:
        print(f"Found {len(accounts)} accounts in token cache")
        try:
            result = client.acquire_token_silent(SCOPES, account=accounts[0])
            if result and "access_token" in result:
                print("Successfully acquired token silently")
                auth_state.update({
                    "token": result["access_token"],
                    "expires_at": now + result.get("expires_in", 3600),
                    "is_authenticated": True
                })
                save_token_cache(cache)  # Only save if we got a valid token
                return result["access_token"]
            else:
                print("Silent token acquisition failed")
                if "error" in result:
                    print(f"Error: {result['error']}")
                    print(f"Error description: {result.get('error_description', 'No description')}")
        except Exception as e:
            print(f"Error during silent token acquisition: {e}")

    if auth_state["flow"]:
        print("Attempting to complete device flow")
        try:
            # Check if the flow has expired
            flow_expiry = auth_state["flow"].get("expires_in", 60)  # Default 60 seconds
            flow_start = auth_state["flow"].get("_start_time", 0)
            time_elapsed = now - flow_start
            print(f"Flow started {time_elapsed:.1f} seconds ago, expires in {flow_expiry} seconds")
            
            if time_elapsed > flow_expiry:
                print("Device flow has expired")
                auth_state["flow"] = None
                raise Exception("Device code has expired. Please initiate a new authentication flow.")

            result = client.acquire_token_by_device_flow(auth_state["flow"])
            print(f"Device flow result: {json.dumps(result, indent=2)}")
            
            if result and "access_token" in result:
                print("Successfully acquired token through device flow")
                auth_state.update({
                    "token": result["access_token"],
                    "expires_at": now + result.get("expires_in", 3600),
                    "is_authenticated": True,
                    "flow": None
                })
                save_token_cache(cache)  # Only save if we got a valid token
                return result["access_token"]
            else:
                print("Device flow token acquisition failed")
                if "error" in result:
                    print(f"Error: {result['error']}")
                    print(f"Error description: {result.get('error_description', 'No description')}")
                    if result["error"] == "authorization_pending":
                        print("Authorization is still pending. Please complete the authentication on the device.")
                        return None
                    elif result["error"] == "expired_token":
                        print("Token has expired. Please initiate a new authentication flow.")
                        auth_state["flow"] = None
                        raise Exception("Token has expired. Please initiate a new authentication flow.")
        except Exception as e:
            print(f"Error during device flow token acquisition: {e}")
            auth_state["flow"] = None

    raise Exception("Not authenticated. Please initiate /auth/initiate")


def cleanup_debug_files(keep_count=2):
    """Keep only the most recent debug Excel files."""
    try:
        files = [f for f in os.listdir(BASE_DIR) if f.startswith("debug_excel_") and f.endswith(".xlsx")]
        if len(files) > keep_count:
            files.sort(reverse=True)
            for old in files[keep_count:]:
                os.remove(os.path.join(BASE_DIR, old))
                print(f"Removed old debug file: {old}")
    except Exception as e:
        print(f"Warning: Failed to cleanup debug files: {e}")


def resolve_drive_item_ids(share_url: str):
    """Resolve and cache driveId and itemId from a 1drv.ms sharing URL."""
    token = get_user_token()
    b64 = base64.urlsafe_b64encode(share_url.encode('utf-8')).decode().rstrip('=')
    share_id = f"u!{b64}"
    meta_url = f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem?$select=id,parentReference"
    headers = {"Authorization": f"Bearer {token}"}
    print(f"Resolving drive and item IDs via: {meta_url}")
    resp = requests.get(meta_url, headers=headers)
    resp.raise_for_status()
    data = resp.json()
    drive_id = data["parentReference"]["driveId"]
    item_id = data["id"]
    print(f"Resolved driveId={drive_id}, itemId={item_id}")
    state.update({"drive_id": drive_id, "item_id": item_id})
    return drive_id, item_id


def download_latest_excel(share_url: str):
    """
    Download the latest Excel content from OneDrive via Graph, bypassing caching.
    """
    token = get_user_token()
    if state.get("current_file") != share_url:
        state["current_file"] = share_url
        state["drive_id"], state["item_id"] = resolve_drive_item_ids(share_url)
    
    # Use the drive item endpoint with select parameter to get latest metadata
    meta_url = f"https://graph.microsoft.com/v1.0/drives/{state['drive_id']}/items/{state['item_id']}?select=id,lastModifiedDateTime,size"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    
    # Get latest metadata
    meta_resp = requests.get(meta_url, headers=headers)
    meta_resp.raise_for_status()
    meta_data = meta_resp.json()
    last_modified = meta_data.get("lastModifiedDateTime")
    file_size = meta_data.get("size", 0)
    print(f"File last modified: {last_modified}, size: {file_size} bytes")
    
    if file_size == 0:
        raise Exception("File appears to be empty or inaccessible. Please check file permissions.")
    
    # Download content with proper headers
    url = f"https://graph.microsoft.com/v1.0/drives/{state['drive_id']}/items/{state['item_id']}/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Cache-Control": "no-cache"  # Only use basic cache control
    }
    
    print(f"Downloading latest file from: {url}")
    try:
        # Calculate optimal chunk size based on file size
        # Use 1MB chunks for files up to 10MB, 5MB chunks for larger files
        chunk_size = 1024 * 1024 if file_size <= 10 * 1024 * 1024 else 5 * 1024 * 1024
        print(f"Using chunk size of {chunk_size/1024/1024:.1f}MB for download")
        
        resp = requests.get(url, headers=headers, stream=True)
        resp.raise_for_status()
        
        # Check content type
        content_type = resp.headers.get('content-type', '')
        print(f"Response content type: {content_type}")
        
        # Read content in chunks
        content = bytearray()
        total_downloaded = 0
        start_time = time.time()
        
        for chunk in resp.iter_content(chunk_size=chunk_size):
            if chunk:
                content.extend(chunk)
                total_downloaded += len(chunk)
                elapsed = time.time() - start_time
                speed = total_downloaded / (1024 * 1024 * elapsed) if elapsed > 0 else 0
                print(f"Downloaded {total_downloaded/1024/1024:.1f}MB of {file_size/1024/1024:.1f}MB ({speed:.1f}MB/s)")
        
        if len(content) == 0:
            raise Exception("Received empty file from OneDrive. Please check file permissions and try again.")
        
        if len(content) < 100:  # Basic sanity check for Excel file
            print(f"Warning: Downloaded content seems too small ({len(content)} bytes). Content preview: {content[:100]}")
            raise Exception("Downloaded file appears to be invalid. Please check file permissions and try again.")

        # Save debug copy with timestamp
        ts = int(time.time())
        debug_path = os.path.join(BASE_DIR, f"debug_excel_{ts}.xlsx")
        with open(debug_path, 'wb') as f:
            f.write(content)
        print(f"Downloaded {len(content)} bytes; saved debug copy to {debug_path}")
        cleanup_debug_files()
        return content
    except requests.exceptions.RequestException as e:
        print(f"Error downloading file: {str(e)}")
        if hasattr(e.response, 'text'):
            print(f"Response text: {e.response.text}")
        raise Exception(f"Failed to download file: {str(e)}")


def parse_results_from_excel(file_path=None, file_bytes=None):
    """Parse the 'Final results' sheet into discipline, category and competitor list."""
    category = None
    discipline = None
    wb = load_workbook(filename=BytesIO(file_bytes) if file_bytes else file_path, data_only=True)
    sheet = wb["Final results"]
    
    # Dictionary to track skaters who have performed (have received points)
    skater_points = {}
    
    # Try to find the last skater from the Marks sheet
    last_skater_name = None
    if "Marks" in wb.sheetnames:
        try:
            # Skip the first 6 rows to get to the headers
            marks_df = pd.read_excel(
                BytesIO(file_bytes) if file_bytes else file_path,
                sheet_name="Marks",
                engine="openpyxl", 
                skiprows=6
            )
            
            # Skip the next row which contains the subheadings (Tech, Art, Total, Place)
            marks_header = marks_df.iloc[0]
            marks_subheader = marks_df.iloc[1]
            marks_df = marks_df.iloc[2:]  # Skip both header rows
            
            # Drop rows with all NaN
            marks_df = marks_df.dropna(how='all')
            
            # Find the column indexes for each judge's Total score
            judge_total_cols = []
            
            # Identify main column headers for judges
            judge_cols = []
            for col in marks_df.columns:
                if 'Judge' in str(col):
                    judge_cols.append(col)
            
            # Find corresponding Total columns for each judge
            for j, judge_col in enumerate(judge_cols):
                col_idx = list(marks_df.columns).index(judge_col)
                # The "Total" column is 2 columns after the main judge column (Tech, Art, Total, Place)
                if col_idx + 2 < len(marks_df.columns):
                    judge_total_cols.append(marks_df.columns[col_idx + 2])
            
            # Find the name column
            name_col = None
            for col in marks_df.columns:
                if 'Name' in str(col):
                    name_col = col
                    break
            
            # Find the last skater who has non-zero Total scores from all three judges
            if name_col and len(judge_total_cols) == 3:
                last_skater_row = None
                
                for idx, row in marks_df.iterrows():
                    skater_name = str(row[name_col]).strip() if pd.notna(row[name_col]) else ""
                    if not skater_name:
                        continue
                        
                    # Check if all three judges have assigned non-zero Total scores
                    has_all_scores = all(
                        pd.notna(row[col]) and float(row[col]) > 0 
                        for col in judge_total_cols if col in row
                    )
                    
                    # Store the scores for each skater - even if not all judges have scored
                    total_points = [float(row[col]) if pd.notna(row[col]) else 0 for col in judge_total_cols if col in row]
                    skater_points[skater_name] = sum(total_points)
                    
                    if has_all_scores:
                        last_skater_row = row
                
                # Get the name of the last skater if found
                if last_skater_row is not None and name_col in last_skater_row:
                    last_skater_name = str(last_skater_row[name_col]).strip()
                    print(f"Found last skater: {last_skater_name}")
                else:
                    print("No skater with complete scores found")
                    
            print(f"Skaters with points: {len([k for k,v in skater_points.items() if v > 0])}")
        except Exception as e:
            print(f"Error processing Marks sheet: {e}")
            # Continue with the rest of the function even if Marks sheet processing fails
    
    # Find Category line
    for i in range(1, 10):
        for j in range(1, 6):
            v = sheet.cell(row=i, column=j).value
            if isinstance(v, str) and v.strip().startswith("Category"):
                category = v.strip().split(" ", 1)[1].strip()
                # Get discipline from the cell above if it exists
                if i > 1:
                    discipline_cell = sheet.cell(row=i-1, column=j).value
                    if discipline_cell and isinstance(discipline_cell, str):
                        discipline = discipline_cell.strip()
                break
        if category:
            break

    df = pd.read_excel(BytesIO(file_bytes) if file_bytes else file_path,
                       sheet_name="Final results", engine="openpyxl", skiprows=6)
    df = df.dropna(how="all")
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    required = ["Name", "Rank", "Judge 1", "Judge 2", "Judge 3"]
    df = df.dropna(subset=required)

    competitors = []
    for _, r in df.iterrows():
        try:
            # Convert numeric values to integers, handling NaN values
            rank = int(float(r["Rank"])) if pd.notna(r["Rank"]) else 0
            penalty = int(float(r.get("PEN", 0))) if pd.notna(r.get("PEN")) else 0
            judge1 = int(float(r["Judge 1"])) if pd.notna(r["Judge 1"]) else 0
            judge2 = int(float(r["Judge 2"])) if pd.notna(r["Judge 2"]) else 0
            judge3 = int(float(r["Judge 3"])) if pd.notna(r["Judge 3"]) else 0
            
            skater_name = str(r["Name"]).strip()
            
            # Skip skaters who have zero total points (haven't performed yet)
            if skater_name in skater_points and skater_points[skater_name] == 0:
                print(f"Skipping {skater_name} - has not performed yet")
                continue
            
            # Check if this competitor is the last skater
            is_last_skater = False
            if last_skater_name and skater_name == last_skater_name:
                is_last_skater = True
            
            competitors.append({
                "rank": rank,
                "name": skater_name,
                "team": str(r.get("Team", "") or ""),
                "country": str(r.get("Ctry", "") or ""),
                "penalty": penalty,
                "judge1": judge1,
                "judge2": judge2,
                "judge3": judge3,
                "remark": str(r.get("Remark", "") or ""),
                "last_skater": is_last_skater  # Add the last_skater flag
            })
        except (ValueError, TypeError) as e:
            print(f"Error processing row: {e}")
            continue

    print(f"Total competitors after filtering: {len(competitors)}")
    return discipline, category, competitors


async def broadcast_to_operators(msg: dict):
    for ws in operator_connections[:]:
        try:
            await ws.send_json(msg)
        except:
            operator_connections.remove(ws)


async def broadcast_to_public(msg: dict):
    for ws in public_connections[:]:
        try:
            await ws.send_json(msg)
        except:
            public_connections.remove(ws)


@app.post("/auth/initiate")
async def initiate_auth():
    """Initiate MSAL device code flow."""
    try:
        # Validate app configuration
        if not secrets.get("microsoft", {}).get("client_id"):
            error_msg = "Client ID is not configured in secrets.json"
            print(error_msg)
            return JSONResponse(status_code=500, content={"error": error_msg})

        client_id = secrets["microsoft"]["client_id"]
        print(f"Using client ID: {client_id}")
        print(f"Using authority: {AUTHORITY}")
        print(f"Requested scopes: {SCOPES}")

        # Validate client ID format
        if not client_id or len(client_id) != 36:
            error_msg = f"Invalid client ID format: {client_id}"
            print(error_msg)
            return JSONResponse(status_code=500, content={"error": error_msg})

        cache = load_token_cache() or msal.SerializableTokenCache()
        print("Creating MSAL PublicClientApplication...")
        client = msal.PublicClientApplication(
            client_id,
            authority=AUTHORITY,
            token_cache=cache
        )
        
        # Clear any existing flow and token
        auth_state.update({
            "flow": None,
            "token": None,
            "expires_at": 0,
            "is_authenticated": False
        })
        
        print("Initiating device flow...")
        try:
            flow = client.initiate_device_flow(scopes=SCOPES)
            print(f"Device flow response: {json.dumps(flow, indent=2)}")
        except Exception as e:
            print(f"Error initiating device flow: {str(e)}")
            raise
        
        if "error" in flow:
            error_msg = f"Error initiating device flow: {flow['error']}"
            if "error_description" in flow:
                error_msg += f"\nDescription: {flow['error_description']}"
            print(error_msg)
            return JSONResponse(status_code=400, content={"error": error_msg})
        
        # Override the default expiration time to 60 seconds
        flow["expires_in"] = 60
        
        # Add start time to flow for expiry checking
        flow["_start_time"] = time.time()
        print(f"Flow start time: {flow['_start_time']}")
        print(f"Flow expiry: {flow.get('expires_in', 60)} seconds")
            
        auth_state.update({"flow": flow, "is_authenticated": False})
        save_token_cache(cache)
        
        print(f"Device flow initiated successfully")
        print(f"User code: {flow['user_code']}")
        print(f"Message: {flow['message']}")
        print(f"Expires in: {flow.get('expires_in', 60)} seconds")
        print(f"Interval: {flow.get('interval', 5)} seconds")
        
        return {
            "message": flow["message"],
            "user_code": flow["user_code"],
            "verification_url": "https://microsoft.com/devicelogin",
            "expires_in": flow.get("expires_in", 60),  # Default to 60 seconds if not specified
            "interval": flow.get("interval", 5)  # Default to 5 seconds if not specified
        }
    except Exception as e:
        error_msg = f"Failed to initiate authentication: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})


@app.get("/auth/status")
async def get_auth_status():
    """Check authentication status."""
    try:
        print("Checking authentication status...")
        token_source = "memory"  # Default source
        try:
            token = get_user_token()
            if token is None:  # Authorization still pending
                flow = auth_state.get("flow")
                if flow:
                    return {
                        "is_authenticated": False,
                        "message": "Authorization pending. Please complete the authentication on the device.",
                        "user_code": flow["user_code"],
                        "verification_url": "https://microsoft.com/devicelogin",
                        "expires_in": flow.get("expires_in", 60),
                        "interval": flow.get("interval", 5)
                    }
        except Exception as e:
            print(f"Token acquisition failed: {str(e)}")
            flow = auth_state.get("flow")
            if flow:
                return {
                    "is_authenticated": False,
                    "message": str(e),
                    "user_code": flow["user_code"],
                    "verification_url": "https://microsoft.com/devicelogin",
                    "expires_in": flow.get("expires_in", 60),
                    "interval": flow.get("interval", 5)
                }
            return {"is_authenticated": False, "message": str(e)}
        
        # Get account information from auth_state if available
        if auth_state.get("is_authenticated"):
            # Check if token was loaded from cache
            if os.path.exists(TOKEN_CACHE_FILE):
                token_source = "cache"
                print("Token was loaded from cache file")
            else:
                print("Token was acquired through device flow")
            
            # Try to get account info from MSAL
            cache = load_token_cache() or msal.SerializableTokenCache()
            client = msal.PublicClientApplication(
                secrets["microsoft"]["client_id"],
                authority=AUTHORITY,
                token_cache=cache
            )
            
            accounts = client.get_accounts()
            print(f"Found {len(accounts)} accounts in token cache")
            
            if accounts:
                account = accounts[0]
                print(f"Raw account data: {json.dumps(account, indent=2)}")
                account_info = {
                    "username": account.get("username", "Unknown"),
                    "name": account.get("name", "Unknown"),
                    "environment": account.get("environment", "Unknown"),
                    "home_account_id": account.get("home_account_id", "Unknown"),
                    "local_account_id": account.get("local_account_id", "Unknown")
                }
                print(f"Processed account info: {json.dumps(account_info, indent=2)}")
                # Update auth_state with the latest account info
                auth_state["account_info"] = account_info
            else:
                print("No accounts found in token cache")
                account_info = auth_state.get("account_info", {})
            
            return {
                "is_authenticated": True,
                "token_expires_at": auth_state["expires_at"],
                "token_source": token_source,
                "account": account_info,
                "message": "Successfully authenticated"
            }
        
        return {
            "is_authenticated": True,
            "token_expires_at": auth_state["expires_at"],
            "token_source": token_source,
            "account": auth_state.get("account_info", {}),
            "message": "Successfully authenticated"
        }
    except Exception as e:
        print(f"Authentication status check failed: {str(e)}")
        return {"is_authenticated": False, "message": str(e)}


async def check_file_updates():
    """
    Background task that periodically checks if the Excel file has been updated.
    If updates are detected, it automatically refreshes the data and notifies clients.
    """
    while True:
        try:
            # Only check if auto-refresh is enabled and we have a current file
            if state["auto_refresh_enabled"] and state["current_file"]:
                now = time.time()
                # Only check every auto_refresh_interval seconds
                if now - state["last_file_check"] >= state["auto_refresh_interval"]:
                    print(f"Checking for file updates at {datetime.now().isoformat()}")
                    state["last_file_check"] = now
                    
                    # Get file metadata
                    token = get_user_token()
                    if "1drv.ms" in state["current_file"] and state["drive_id"] and state["item_id"]:
                        meta_url = f"https://graph.microsoft.com/v1.0/drives/{state['drive_id']}/items/{state['item_id']}?select=id,lastModifiedDateTime,size"
                        headers = {
                            "Authorization": f"Bearer {token}",
                            "Accept": "application/json",
                            "Cache-Control": "no-cache"
                        }
                        
                        meta_resp = requests.get(meta_url, headers=headers)
                        meta_resp.raise_for_status()
                        meta_data = meta_resp.json()
                        last_modified = meta_data.get("lastModifiedDateTime")
                        
                        # Check if file has been modified
                        if state["last_modified"] != last_modified:
                            print(f"File has been updated: {last_modified}")
                            print(f"Previous modification: {state['last_modified']}")
                            
                            # Update last_modified timestamp
                            state["last_modified"] = last_modified
                            
                            # Refresh data from the updated file
                            try:
                                content = download_latest_excel(state["current_file"])
                                discipline, category, comps = parse_results_from_excel(file_bytes=content)
                                
                                # Update state with new data
                                state["live"]["category"], state["live"]["discipline"], state["live"]["competitors"] = category, discipline, comps
                                
                                # Notify all connected clients
                                print(f"Broadcasting update to {len(operator_connections)} operators")
                                await broadcast_to_operators({
                                    "type": "live_update", 
                                    "data": state["live"],
                                    "auto_refreshed": True,
                                    "timestamp": datetime.now().isoformat()
                                })
                                
                                print(f"Auto-refresh successful: {len(comps)} competitors")
                            except Exception as e:
                                print(f"Error during auto-refresh: {str(e)}")
                        else:
                            print(f"No file changes detected.")
            
            # Sleep before next check
            await asyncio.sleep(5)  # Check every 5 seconds
            
        except Exception as e:
            print(f"Error in background task: {str(e)}")
            await asyncio.sleep(10)  # Wait longer if there's an error

@app.on_event("startup")
async def startup_event():
    """Start background tasks when the app starts."""
    print("Starting background tasks...")
    
    # Start file update checker
    task = asyncio.create_task(check_file_updates())
    background_tasks.add(task)
    task.add_done_callback(background_tasks.discard)
    
    print("Background tasks started successfully.")

@app.get("/auto_refresh/status")
async def get_auto_refresh_status():
    """Get the current auto-refresh settings."""
    return {
        "enabled": state["auto_refresh_enabled"],
        "interval": state["auto_refresh_interval"],
        "last_check": state["last_file_check"],
        "last_modified": state["last_modified"]
    }

@app.post("/auto_refresh/settings")
async def update_auto_refresh_settings(settings: dict):
    """Update auto-refresh settings."""
    if "enabled" in settings:
        state["auto_refresh_enabled"] = bool(settings["enabled"])
    
    if "interval" in settings:
        interval = int(settings["interval"])
        if interval >= 5:  # Minimum 5 seconds
            state["auto_refresh_interval"] = interval
    
    return {
        "enabled": state["auto_refresh_enabled"],
        "interval": state["auto_refresh_interval"]
    }

@app.post("/load_excel")
async def load_excel(source: dict):
    url = source.get("url")
    path = source.get("path")
    try:
        if url and "1drv.ms" in url:
            print(f"Loading Excel from OneDrive URL: {url}")
            try:
                content = download_latest_excel(url)
            except Exception as e:
                error_msg = f"Failed to download from OneDrive: {str(e)}"
                print(error_msg)
                return JSONResponse(status_code=400, content={"error": error_msg})
        elif url:
            print(f"Loading Excel from URL: {url}")
            r = requests.get(url)
            r.raise_for_status()
            content = r.content
        else:
            content = None
            print("No URL provided, using default data file")
        
        if content:
            print("Parsing Excel content")
            discipline, category, comps = parse_results_from_excel(file_bytes=content)
            print(f"Successfully parsed {len(comps)} competitors")
        else:
            print("Using default data file")
            discipline, category, comps = parse_results_from_excel(file_path=DEFAULT_DATA_FILE)
        
        # Update state with last_modified timestamp for auto-refresh
        if url and "1drv.ms" in url and state["drive_id"] and state["item_id"]:
            try:
                token = get_user_token()
                meta_url = f"https://graph.microsoft.com/v1.0/drives/{state['drive_id']}/items/{state['item_id']}?select=lastModifiedDateTime"
                headers = {
                    "Authorization": f"Bearer {token}",
                    "Accept": "application/json"
                }
                meta_resp = requests.get(meta_url, headers=headers)
                meta_resp.raise_for_status()
                state["last_modified"] = meta_resp.json().get("lastModifiedDateTime")
                print(f"File last modified timestamp: {state['last_modified']}")
            except Exception as e:
                print(f"Warning: Could not get last_modified timestamp: {str(e)}")
        
        state["live"] = {"category": category, "discipline": discipline, "competitors": comps, "category_complete": False}
        await broadcast_to_operators({"type": "live_update", "data": state["live"]})
        return {"status": "ok", "category": category, "count": len(comps)}
    except Exception as e:
        error_msg = f"Error processing file: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=400, content={"error": error_msg})


@app.post("/refresh_data")  
async def refresh_data():
    if not state.get("current_file"):
        return JSONResponse(status_code=400, content={"error": "No data source loaded."})
    try:
        print(f"Refreshing data from: {state['current_file']}")
        if "1drv.ms" in state["current_file"]:
            print("Using OneDrive download for refresh")
            content = download_latest_excel(state["current_file"])
        else:
            print("Using regular download for refresh")
            headers = {
                "Cache-Control": "no-cache, no-store, must-revalidate",
                "Pragma": "no-cache",
                "Expires": "0",
                "If-None-Match": "*",
                "If-Modified-Since": "0"
            }
            r = requests.get(state["current_file"], headers=headers)
            r.raise_for_status()
            content = r.content
        print("Parsing refreshed Excel content")
        discipline, category, comps = parse_results_from_excel(file_bytes=content)
        print(f"Parsed {len(comps)} competitors from refreshed data")
        state["live"] = {"category": category, "discipline": discipline, "competitors": comps, "category_complete": False}
        await broadcast_to_operators({"type": "live_update", "data": state["live"]})
        return {"status": "ok", "updated_count": len(comps)}
    except Exception as e:
        print(f"Error in refresh_data: {e}")
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.post("/publish")
async def publish_public():
    if not state["live"]["competitors"]:
        return JSONResponse(status_code=400, content={"error": "No results to publish."})
    limit = config.get("style", {}).get("publicDisplayLimit")
    state["public"] = {
        "category": state["live"]["category"],
        "discipline": state["live"]["discipline"],
        "competitors": state["live"]["competitors"][:limit] if limit else state["live"]["competitors"],
        "category_complete": state["live"]["category_complete"],
        "message": "",
        "display_mode": "results"  # Set to results mode when publishing
    }
    await broadcast_to_public({"type": "public_update", "data": state["public"]})
    await broadcast_to_operators({"type": "public_update", "data": state["public"]})
    return {"status": "ok", "published_count": len(state["public"]["competitors"])}


@app.post("/mark_complete")
async def mark_complete(data: dict = None):
    # Default to True if no data is provided, otherwise use the provided value
    complete_status = True if data is None else data.get("category_complete", True)
    state["live"]["category_complete"] = complete_status
    await broadcast_to_operators({"type": "live_update", "data": state["live"]})
    return {"status": "ok", "category_complete": complete_status}


@app.post("/display_message")
async def display_message(msg: dict):
    state["public"]["message"] = msg.get("message", "")
    state["public"]["display_mode"] = "message"  # Set to message mode
    await broadcast_to_public({"type": "public_update", "data": state["public"]})
    await broadcast_to_operators({"type": "public_update", "data": state["public"]})
    return {"status": "ok", "message": state["public"]["message"], "display_mode": "message"}


@app.post("/switch_display_mode")
async def switch_display_mode(data: dict):
    mode = data.get("mode", "results")
    if mode not in ["results", "message"]:
        return JSONResponse(status_code=400, content={"error": "Invalid display mode. Must be 'results' or 'message'."})
    
    state["public"]["display_mode"] = mode
    await broadcast_to_public({"type": "public_update", "data": state["public"]})
    await broadcast_to_operators({"type": "public_update", "data": state["public"]})
    return {"status": "ok", "display_mode": mode}


@app.post("/upload_background")
async def upload_background(file: UploadFile = File(...)):
    # File will be uploaded to frontend/public/backgrounds/ instead of backend/static/
    filename = file.filename
    
    # Create the backgrounds directory if it doesn't exist
    frontend_bg_dir = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend", "public", "backgrounds")
    os.makedirs(frontend_bg_dir, exist_ok=True)
    
    # Save the file to the frontend directory
    path = os.path.join(frontend_bg_dir, filename)
    with open(path, "wb") as f:
        f.write(await file.read())
    
    # Update the URL to use the frontend path format - no need for server prefix
    # This will be resolved relative to the frontend application
    state["background_url"] = f"/backgrounds/{filename}"
    
    await broadcast_to_public({"type": "background_update", "url": state["background_url"]})
    await broadcast_to_operators({"type": "background_update", "url": state["background_url"]})
    return {"status": "ok", "background_url": state["background_url"]}


@app.get("/config")
async def get_config():
    return {
        "style": config.get("style", {}),
        "backgroundUrl": state["background_url"],
        "defaultExcelUrl": config.get("default_excel_url", ""),
        "worldSkateRankingsUrl": config.get("worldSkateRankingsUrl", "https://app-69b8883b-99d4-4935-9b2b-704880862424.cleverapps.io")
    }


@app.websocket("/ws/operator")
async def operator_ws(ws: WebSocket):
    await ws.accept()
    operator_connections.append(ws)
    if state["live"]["competitors"]:
        await ws.send_json({"type": "live_update", "data": state["live"]})
    if state["public"]["competitors"]:
        await ws.send_json({"type": "public_update", "data": state["public"]})
    if state["background_url"]:
        await ws.send_json({"type": "background_update", "url": state["background_url"]})
    try:
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        operator_connections.remove(ws)


@app.websocket("/ws/public")
async def public_ws(ws: WebSocket):
    await ws.accept()
    public_connections.append(ws)
    if state["public"]["competitors"]:
        await ws.send_json({"type": "public_update", "data": state["public"]})
    if state["background_url"]:
        await ws.send_json({"type": "background_update", "url": state["background_url"]})
    try:
        while True:
            await ws.receive_text()
    except WebSocketDisconnect:
        public_connections.remove(ws)


# Add new endpoints for registration handling

@app.post("/google/auth/initiate")
async def google_auth_initiate():
    """Initiate Google OAuth2 flow."""
    try:
        result = initiate_auth_flow()
        if "error" in result:
            return JSONResponse(status_code=500, content={"error": result["error"]})
        return result
    except Exception as e:
        error_msg = f"Failed to initiate Google authentication: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.get("/api/google/auth/callback")
async def api_google_auth_callback(code: str = Query(None), state: str = Query(None)):
    """API endpoint for Google OAuth callback."""
    if not code:
        return JSONResponse(status_code=400, content={"error": "No authentication code provided"})
    
    try:
        result = complete_auth_flow(code)
        if "error" in result:
            return JSONResponse(status_code=500, content={"error": result["error"]})
        return {"success": True, "message": "Authentication successful"}
    except Exception as e:
        error_msg = f"Failed to complete Google authentication: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.get("/google/auth/status")
async def google_auth_status():
    """Check Google authentication status."""
    try:
        from google_sheets import auth_state
        
        credentials = get_credentials()
        if credentials:
            # Use the user_info from auth_state if available
            user_info = auth_state.get("user_info", {})
            email = user_info.get("email", "Unknown")
            
            return {
                "is_authenticated": True,
                "email": email,
                "user_info": user_info
            }
        return {
            "is_authenticated": False,
            "message": "Not authenticated with Google"
        }
    except Exception as e:
        error_msg = f"Failed to check Google authentication status: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.post("/registration/load")
async def load_registration(source: dict):
    """Load registration data from Google Sheets."""
    url = source.get("url")
    if not url:
        return JSONResponse(status_code=400, content={"error": "No URL provided"})
    
    try:
        # Fetch data from Google Sheets
        sheet_data = fetch_spreadsheet_data(url)
        if "error" in sheet_data:
            return JSONResponse(status_code=400, content={"error": sheet_data["error"]})
        
        # Parse registration data
        result = parse_registration_data(sheet_data["csv_data"])
        if "error" in result:
            return JSONResponse(status_code=400, content={"error": result["error"]})
        
        # Update registration state
        reg_state["current_sheet_url"] = url
        reg_state["disciplines"] = result["disciplines"]
        reg_state["skaters"] = result["skaters"]
        
        return {
            "success": True,
            "message": f"Successfully loaded {len(result['skaters'])} skaters with {len(result['disciplines'])} disciplines",
            "disciplines": result["disciplines"],
            "skaters_count": len(result["skaters"]),
            "document_title": sheet_data.get("document_title", "Google Sheet")
        }
    except Exception as e:
        error_msg = f"Failed to load registration data: {str(e)}"
        print(error_msg)
        return JSONResponse(status_code=500, content={"error": error_msg})

@app.get("/registration/disciplines")
async def get_disciplines():
    """Get list of disciplines from registration data."""
    return {"disciplines": reg_state["disciplines"]}

@app.get("/registration/skaters")
async def get_skaters(discipline: str = None, sex: str = None):
    """Get list of skaters from registration data, with optional filtering."""
    if not reg_state["skaters"]:
        return {"error": "No registration data loaded"}
    
    skaters = reg_state["skaters"]
    
    # Apply discipline filter
    if discipline:
        skaters = [s for s in skaters if discipline in s["disciplines"]]
    
    # Apply sex filter
    if sex and sex.upper() in ["M", "F"]:
        skaters = [s for s in skaters if s["sex"] == sex.upper()]
    
    # Clean NaN values from skaters data
    clean_skaters = []
    for skater in skaters:
        clean_skater = {}
        for key, value in skater.items():
            # Check if value is NaN (not equal to itself)
            if isinstance(value, float) and value != value:  # NaN check
                clean_skater[key] = None
            else:
                clean_skater[key] = value
        clean_skaters.append(clean_skater)
    
    return {"skaters": clean_skaters, "count": len(clean_skaters)}

# Create an API endpoint to check if frontend is available
@app.get("/api/check-frontend")
async def check_frontend():
    """Check if the frontend build is available."""
    if FRONTEND_DIR and os.path.exists(os.path.join(FRONTEND_DIR, "index.html")):
        return {"frontend_available": True, "frontend_dir": FRONTEND_DIR}
    else:
        # List potential directories we checked
        potential_dirs = [
            os.path.join(os.path.dirname(__file__), "..", "frontend", "build"),
            os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"),
            os.path.join(os.path.dirname(__file__), "frontend", "build"),
            os.path.join(os.path.dirname(__file__), "frontend", "dist")
        ]
        
        return {
            "frontend_available": False,
            "checked_paths": potential_dirs,
            "error": "Frontend build not found. Please run the build script first."
        }

# Add a special route for Google Auth callback from OAuth
@app.get("/reg")
async def handle_reg_route(request: Request, code: str = Query(None), state: str = Query(None)):
    """Special handler for /reg route with Google Auth callback."""
    print(f"Handling /reg route with code: {code is not None}, state: {state is not None}")
    
    # If we have code and state, this is a Google Auth callback
    if code and state:
        try:
            # Process the authentication
            result = complete_auth_flow(code)
            if "error" in result:
                print(f"Auth error: {result['error']}")
            else:
                print("Authentication successful!")
                
            # Now redirect to the frontend app
            if FRONTEND_DIR and os.path.exists(os.path.join(FRONTEND_DIR, "index.html")):
                print(f"Redirecting to frontend app at {FRONTEND_DIR}")
                # Set a query param to indicate successful auth
                return RedirectResponse(url="/reg?auth_success=true")
        except Exception as e:
            print(f"Error in Google Auth callback: {str(e)}")
    
    # For a regular request to /reg, serve the React app
    if FRONTEND_DIR and os.path.exists(os.path.join(FRONTEND_DIR, "index.html")):
        print(f"Serving index.html from {FRONTEND_DIR} for /reg route")
        return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))
    
    # If frontend not found, return a helpful error
    return JSONResponse(
        status_code=404,
        content={
            "detail": "Frontend not found",
            "message": "React app build files not found. Please run the build script."
        }
    )

# Rankings API Endpoints
@app.get("/rankings/info")
async def get_rankings_info():
    """Get information about the current rankings status"""
    try:
        # Base URL of the World Skate rankings application from config
        base_url = config.get("worldSkateRankingsUrl", "https://app-69b8883b-99d4-4935-9b2b-704880862424.cleverapps.io")
        
        # Find the latest rankings directory by scanning folder names
        rankings_dir = "rankings"
        os.makedirs(rankings_dir, exist_ok=True)
        
        # Find the latest rankings directory by scanning folder names
        latest_rankings_path = get_latest_rankings_folder(rankings_dir)
        latest_date = None
        available_disciplines = []
        
        # Check if we found a latest rankings directory
        if latest_rankings_path and os.path.exists(latest_rankings_path):
            # Get the folder name (which will be in "YYYY-MM_Month" format)
            latest_date = os.path.basename(latest_rankings_path)
            
            # List all available disciplines from the latest rankings directory
            for file in os.listdir(latest_rankings_path):
                if file.endswith('.csv'):
                    # Remove the .csv extension
                    discipline = file[:-4]
                    available_disciplines.append(discipline)
        
        # Check the World Skate website for the most recent rankings date
        external_latest_date = None
        newer_available = False
        
        try:
            response = requests.get(base_url, timeout=10)
            if response.status_code == 200:
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Extract the latest ranking date from the first archive link
                archives_div = soup.find('div', class_='left-filters')
                latest_link = archives_div.find('a') if archives_div else None
                if latest_link:
                    # Convert from YYYY-MM-DD to YYYY-MM_Month format
                    external_latest_date_raw = latest_link.get_text(strip=True)
                    external_latest_date = format_date_for_folder(external_latest_date_raw)
                    
                    # Check if a newer version is available
                    newer_available = latest_date != external_latest_date
        except Exception as e:
            print(f"Error checking for latest rankings: {e}")
        
        return {
            "latest_date": latest_date,
            "external_latest_date": external_latest_date,
            "newer_available": newer_available,
            "available_disciplines": sorted(available_disciplines)
        }
    except Exception as e:
        print(f"Error getting rankings info: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get rankings info: {str(e)}"}
        )

@app.post("/rankings/update")
async def update_rankings(background_tasks: BackgroundTasks):
    """Trigger a rankings update in the background"""
    try:
        # Use background tasks to run the rankings update without blocking
        background_tasks.add_task(fetch_rankings)
        
        return {"status": "updating", "message": "Rankings update has been initiated"}
    except Exception as e:
        print(f"Error starting rankings update: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to start rankings update: {str(e)}"}
        )

@app.get("/api/rankings/progress")
async def get_rankings_progress():
    """Get the current progress of the rankings download"""
    try:
        progress = get_download_progress()
        return progress
    except Exception as e:
        print(f"Error getting rankings progress: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get rankings progress: {str(e)}"}
        )

# Add API versions of the rankings endpoints
@app.get("/api/rankings/info")
async def api_rankings_info():
    """API version of the rankings info endpoint"""
    return await get_rankings_info()

@app.post("/api/rankings/update")
async def api_rankings_update(background_tasks: BackgroundTasks):
    """API version of the rankings update endpoint"""
    return await update_rankings(background_tasks)

# Make sure this endpoint is properly registered
@app.get("/api/rankings/download-zip", response_class=Response)
async def api_rankings_download_zip():
    """API version of the download rankings zip endpoint"""
    print("API Download ZIP endpoint called!")  # Debug print
    return await download_rankings_zip()

async def download_rankings_zip():
    """Create and return a zip file of the latest rankings"""
    print("Download ZIP endpoint called!")  # Debug print
    try:
        import zipfile
        import io

        # Find the latest rankings directory
        rankings_dir = "rankings"
        latest_rankings_path = get_latest_rankings_folder(rankings_dir)
        
        if not latest_rankings_path or not os.path.exists(latest_rankings_path):
            print(f"No rankings folder found at: {rankings_dir}")  # Debug print
            return JSONResponse(
                status_code=404,
                content={"error": "No rankings data found"}
            )
        
        print(f"Creating zip from folder: {latest_rankings_path}")  # Debug print
        
        # Create a BytesIO object to hold the zip file
        zip_buffer = io.BytesIO()
        
        # Create a zip file in the buffer
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # Get the folder name (e.g., "2023-06_Jun")
            folder_name = os.path.basename(latest_rankings_path)
            
            # Add each CSV file to the zip
            files_added = 0  # Debug counter
            for file_name in os.listdir(latest_rankings_path):
                if file_name.endswith('.csv'):
                    file_path = os.path.join(latest_rankings_path, file_name)
                    # Add the file to the zip with a path inside the zip
                    zip_file.write(file_path, os.path.join(folder_name, file_name))
                    files_added += 1  # Debug counter
            
            print(f"Added {files_added} files to zip")  # Debug print
        
        # Seek to the beginning of the buffer
        zip_buffer.seek(0)
        
        print("Returning zip file")  # Debug print
        
        # Return the zip file
        return Response(
            content=zip_buffer.getvalue(),
            media_type="application/zip",
            headers={
                "Content-Disposition": f"attachment; filename=rankings_{folder_name}.zip"
            }
        )
    except Exception as e:
        print(f"Error creating zip file: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to create zip file: {str(e)}"}
        )

@app.get("/api/rankings/table-metadata")
async def get_table_metadata():
    """Get the table metadata for World Skate rankings"""
    try:
        # Find the latest rankings directory
        rankings_dir = "rankings"
        latest_rankings_path = get_latest_rankings_folder(rankings_dir)
        
        if not latest_rankings_path or not os.path.exists(latest_rankings_path):
            return JSONResponse(
                status_code=404,
                content={"error": "No rankings data found"}
            )
        
        # Read the table metadata file
        metadata_file = os.path.join(latest_rankings_path, "table_metadata.json")
        if not os.path.exists(metadata_file):
            return JSONResponse(
                status_code=404,
                content={"error": "Table metadata not found"}
            )
        
        with open(metadata_file, 'r') as f:
            metadata = json.load(f)
        
        return metadata
    except Exception as e:
        print(f"Error getting table metadata: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get table metadata: {str(e)}"}
        )

@app.get("/api/rankings/all/combined")
async def get_all_rankings():
    """Get all rankings in a single combined JSON response"""
    print(f"[DEBUG] Processing /api/rankings/all/combined request")
    try:
        # Find the latest rankings directory
        rankings_dir = "rankings"
        latest_rankings_path = get_latest_rankings_folder(rankings_dir)
        
        if not latest_rankings_path or not os.path.exists(latest_rankings_path):
            print(f"[DEBUG] No rankings found at path: {rankings_dir}")
            return JSONResponse(
                status_code=404,
                content={"error": "No rankings data found"}
            )
        
        # Dictionary to store all rankings by discipline
        all_rankings = {}
        latest_update = os.path.basename(latest_rankings_path)
        print(f"[DEBUG] Found rankings directory: {latest_update}")
        
        # Read all CSV files in the directory
        file_count = 0
        for file in os.listdir(latest_rankings_path):
            if file.endswith('.csv'):
                file_count += 1
                # Get discipline name from filename (without extension)
                discipline = file[:-4]
                file_path = os.path.join(latest_rankings_path, file)
                
                try:
                    # Read the CSV file
                    df = pd.read_csv(file_path, quoting=csv.QUOTE_ALL)
                    
                    # Create a mapping of World Skate IDs to rankings
                    id_to_rank = {}
                    for _, row in df.iterrows():
                        ws_id = str(row["ID"]).strip()
                        if ws_id:  # Only process entries with valid World Skate IDs
                            id_to_rank[ws_id] = {
                                "rank": row["Rank"],
                                "name": row["Name"],
                                "country": row["Nat."],
                                "points": row["Best"]
                            }
                    
                    # Add to the overall dictionary
                    all_rankings[discipline] = id_to_rank
                    
                except Exception as e:
                    print(f"[DEBUG] Error reading {file}: {e}")
                    continue
        
        print(f"[DEBUG] Successfully processed {file_count} ranking files, found {len(all_rankings)} disciplines")
        return {
            "latest_update": latest_update,
            "rankings": all_rankings
        }
    except Exception as e:
        print(f"[DEBUG] Error getting all rankings: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get all rankings: {str(e)}"}
        )

@app.get("/api/rankings/{discipline}")
async def get_discipline_rankings(discipline: str):
    """Get the rankings data for a specific discipline"""
    try:
        # Get the file path for the specified discipline
        file_path = get_discipline_file_path(discipline)
        
        if not file_path or not os.path.exists(file_path):
            return JSONResponse(
                status_code=404,
                content={"error": f"Discipline '{discipline}' not found in rankings data"}
            )
        
        # Read the CSV file using pandas
        import pandas as pd
        df = pd.read_csv(file_path, quoting=csv.QUOTE_ALL)
        
        # Convert to list of dictionaries for JSON response
        rankings = []
        for _, row in df.iterrows():
            rankings.append({
                "rank": row["Rank"],
                "name": row["Name"],
                "country": row["Nat."],
                "world_skate_id": row["ID"],
                "best_points": row["Best"]
            })
        
        return {"rankings": rankings}
    except Exception as e:
        print(f"Error getting discipline rankings: {e}")
        return JSONResponse(
            status_code=500,
            content={"error": f"Failed to get discipline rankings: {str(e)}"}
        )

# Modify the catch-all route to always serve index.html for frontend routes
@app.get("/{catch_all:path}")
async def serve_react_app(request: Request, catch_all: str):
    path = request.url.path
    print(f"[DEBUG] Catch-all route for path: {path}")
    
    # Add specific debug logs for rankings routes
    if path.startswith("/api/rankings/"):
        print(f"[DEBUG] Rankings API path detected: {path}")
    elif path.startswith("/api/skater-db/"):
        print(f"[DEBUG] Skater DB API path detected: {path}")
        # Important: Immediately exit the catch-all for skater-db API paths
        # They should be handled by their specific route handlers defined earlier
        return JSONResponse(status_code=404, content={"detail": "API endpoint not found"})
    
    # Skip API routes
    if path.startswith("/api/") or path.startswith("/ws/") or (path.startswith("/rankings/") and not path.startswith("/rankings/info") and not path.startswith("/rankings/update")):
        print(f"[DEBUG] Skipping API route in catch-all: {path}")
        return JSONResponse(status_code=404, content={"detail": "Not Found - API route not defined"})
    
    # List all potential frontend dirs to check
    print(f"[DEBUG] FRONTEND_DIR set to: {FRONTEND_DIR}")
    all_frontend_dirs = [
        os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"),
        os.path.join(os.path.dirname(__file__), "..", "frontend", "build"),
        os.path.join(os.path.dirname(__file__), "frontend", "dist"),
        os.path.join(os.path.dirname(__file__), "frontend", "build")
    ]
    
    # Add FRONTEND_DIR if it's set
    if FRONTEND_DIR:
        all_frontend_dirs.insert(0, FRONTEND_DIR)
    
    # Print out all dirs we'll check
    for i, dir_path in enumerate(all_frontend_dirs):
        exists = os.path.exists(dir_path)
        index_exists = os.path.exists(os.path.join(dir_path, "index.html"))
        print(f"[DEBUG] Frontend dir {i}: {dir_path} - Exists: {exists}, index.html exists: {index_exists}")
    
    # Check if frontend files exist in mounted directory
    if FRONTEND_DIR and os.path.exists(os.path.join(FRONTEND_DIR, "index.html")):
        print(f"[DEBUG] Serving index.html from mounted FRONTEND_DIR: {FRONTEND_DIR}")
        return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))
    
    # Check all potential frontend directories
    for frontend_dir in all_frontend_dirs:
        index_path = os.path.join(frontend_dir, "index.html")
        if os.path.exists(index_path):
            print(f"[DEBUG] Found and serving index.html from: {frontend_dir}")
            return FileResponse(index_path)
    
    # If frontend not found, return a helpful error
    print("[DEBUG] Could not find index.html in any potential frontend directory")
    return JSONResponse(
        status_code=404,
        content={
            "detail": "Frontend not found",
            "message": "React app build files not found. Please run the build_frontend script.",
            "checked_directories": all_frontend_dirs
        }
    )

# Add a test route before any other routes
@app.get("/test")
async def test_static_file():
    """Test endpoint to check if static file serving is working."""
    test_file = os.path.join(os.path.dirname(os.path.abspath(__file__)), "test_page.html")
    if os.path.exists(test_file):
        return FileResponse(test_file)
    return {"error": "Test file not found", "path": test_file}

# Add another test route for React app serving
@app.get("/debug-frontend")
async def debug_frontend_serving():
    """Debug endpoint to check React app serving."""
    results = {
        "frontend_dir": FRONTEND_DIR,
        "frontend_dir_found": FRONTEND_DIR is not None,
        "index_exists": False,
        "static_dir_exists": False,
        "static_dir_path": None,
        "index_content_sample": None
    }
    
    # Check if FRONTEND_DIR exists and has index.html
    if FRONTEND_DIR:
        index_path = os.path.join(FRONTEND_DIR, "index.html")
        static_dir = os.path.join(FRONTEND_DIR, "static")
        
        results["index_exists"] = os.path.exists(index_path)
        results["static_dir_exists"] = os.path.exists(static_dir)
        results["static_dir_path"] = static_dir if os.path.exists(static_dir) else None
        
        # Get a sample of index.html content if it exists
        if results["index_exists"]:
            try:
                with open(index_path, 'r') as f:
                    content = f.read(500)  # Read first 500 chars
                    results["index_content_sample"] = content + "..." if len(content) >= 500 else content
            except Exception as e:
                results["index_content_sample"] = f"Error reading file: {str(e)}"
    
    # Check potential frontend directories if FRONTEND_DIR is not set
    potential_dirs = [
        os.path.join(os.path.dirname(__file__), "..", "frontend", "build"),
        os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"),
        os.path.join(os.path.dirname(__file__), "frontend", "build"),
        os.path.join(os.path.dirname(__file__), "frontend", "dist")
    ]
    
    results["checked_potential_dirs"] = []
    
    for dir_path in potential_dirs:
        dir_info = {
            "path": dir_path,
            "exists": os.path.exists(dir_path),
            "index_exists": os.path.exists(os.path.join(dir_path, "index.html")) if os.path.exists(dir_path) else False
        }
        results["checked_potential_dirs"].append(dir_info)
    
    return results

# Add static files route for frontend assets
@app.get("/static/{rest_of_path:path}")
async def serve_static_files(request: Request, rest_of_path: str):
    print(f"[DEBUG] Handling static file request: /static/{rest_of_path}")
    
    # Try to find the static file in potential frontend directories
    static_path = None
    
    # First check FRONTEND_DIR if set
    if FRONTEND_DIR:
        potential_path = os.path.join(FRONTEND_DIR, "static", rest_of_path)
        if os.path.exists(potential_path) and os.path.isfile(potential_path):
            static_path = potential_path
            print(f"[DEBUG] Found static file in FRONTEND_DIR: {static_path}")
    
    # If not found, check other potential directories
    if not static_path:
        for frontend_dir in [
            os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"),
            os.path.join(os.path.dirname(__file__), "..", "frontend", "build"),
            os.path.join(os.path.dirname(__file__), "frontend", "dist"),
            os.path.join(os.path.dirname(__file__), "frontend", "build")
        ]:
            potential_path = os.path.join(frontend_dir, "static", rest_of_path)
            if os.path.exists(potential_path) and os.path.isfile(potential_path):
                static_path = potential_path
                print(f"[DEBUG] Found static file in alternative dir: {static_path}")
                break
    
    # If found, return the file
    if static_path:
        return FileResponse(static_path)
    
    # Otherwise, return 404
    print(f"[DEBUG] Static file not found: /static/{rest_of_path}")
    return JSONResponse(
        status_code=404,
        content={"detail": f"Static file not found: /static/{rest_of_path}"}
    )

if __name__ == "__main__":
    print("Testing Microsoft Graph API authentication...")
    import asyncio

    async def test_auth():
        status = await get_auth_status()
        if status.get("is_authenticated"):
            print("Authentication test successful!")
        else:
            print("Authentication test failed!")
            print(f"Status: {status.get('message')}")
            exit(1)

    asyncio.run(test_auth())
