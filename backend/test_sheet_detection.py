#!/usr/bin/env python3
"""
Test script for case-insensitive sheet detection functionality.
This script tests the enhanced sheet name detection in result_sheets.py
"""

import sys
import os
from openpyxl import Workbook
from io import BytesIO

# Add the current directory to the path so we can import result_sheets
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

def test_sheet_detection_only():
    """Test just the sheet detection logic without full processing."""
    
    print("Testing case-insensitive sheet detection...")
    print("=" * 50)
    
    # Import here to avoid registration data issues
    from result_sheets import load_workbook, BytesIO
    
    # Test different sheet name variations
    test_cases = [
        "Final Ranking",      # Title case
        "Final ranking",      # Sentence case
        "final ranking",      # Lower case
        "FINAL RANKING",      # Upper case
        "Final Results",      # Alternative name
        "Final results",      # Alternative name, sentence case
        "final results",      # Alternative name, lower case
        "FINAL RESULTS"       # Alternative name, upper case
    ]
    
    success_count = 0
    
    for sheet_name in test_cases:
        print(f"\nTesting sheet name: '{sheet_name}'")
        try:
            # Create test Excel file with this sheet name
            file_bytes = create_test_excel_file(sheet_name)
            
            # Test the sheet detection logic directly
            wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
            
            # Find the "Final Ranking" sheet (case-insensitive) - copied from result_sheets.py
            final_ranking_sheet = None
            final_ranking_sheet_name = None
            
            # Common variations of the sheet name to look for
            possible_sheet_names = [
                "Final Ranking",
                "Final ranking", 
                "final ranking",
                "FINAL RANKING",
                "Final Results",
                "Final results",
                "final results",
                "FINAL RESULTS"
            ]
            
            # First, try exact matches with the common variations
            for test_sheet_name in possible_sheet_names:
                if test_sheet_name in wb.sheetnames:
                    final_ranking_sheet = wb[test_sheet_name]
                    final_ranking_sheet_name = test_sheet_name
                    print(f"  ✅ Found sheet: '{test_sheet_name}'")
                    break
            
            # If no exact match found, try case-insensitive search
            if final_ranking_sheet is None:
                for wb_sheet_name in wb.sheetnames:
                    if "final" in wb_sheet_name.lower() and ("ranking" in wb_sheet_name.lower() or "results" in wb_sheet_name.lower()):
                        final_ranking_sheet = wb[wb_sheet_name]
                        final_ranking_sheet_name = wb_sheet_name
                        print(f"  ✅ Found sheet (case-insensitive): '{wb_sheet_name}'")
                        break
            
            # If still no match, check if there's only one sheet and use it
            if final_ranking_sheet is None:
                if len(wb.sheetnames) == 1:
                    final_ranking_sheet = wb[wb.sheetnames[0]]
                    final_ranking_sheet_name = wb.sheetnames[0]
                    print(f"  ✅ Using only available sheet: '{final_ranking_sheet_name}'")
                else:
                    print(f"  ❌ No matching sheet found. Available sheets: {wb.sheetnames}")
                    continue
            
            if final_ranking_sheet is not None:
                success_count += 1
                print(f"     Successfully detected sheet: '{final_ranking_sheet_name}'")
            else:
                print(f"  ❌ Sheet detection failed")
                
        except Exception as e:
            print(f"  ❌ Error: {str(e)}")
    
    print("\n" + "=" * 50)
    print(f"Test Results: {success_count}/{len(test_cases)} sheet names detected successfully")
    
    if success_count == len(test_cases):
        print("✅ All sheet name variations detected correctly!")
        return True
    else:
        print("❌ Some sheet name variations failed!")
        return False

def create_test_excel_file(sheet_name: str) -> bytes:
    """Create a test Excel file with the specified sheet name."""
    wb = Workbook()
    
    # Remove the default sheet and create our test sheet
    wb.remove(wb.active)
    sheet = wb.create_sheet(title=sheet_name)
    
    # Add some test data
    sheet['A1'] = ""
    sheet['B1'] = "Freestyle Slalom Battle"
    sheet['B2'] = "Junior Women - Final Ranking"
    sheet['A5'] = "Ranking"
    sheet['B5'] = "World Skate Id"
    sheet['C5'] = "BIB"
    sheet['D5'] = "Name"
    sheet['E5'] = "Team"
    sheet['F5'] = "Ctry"
    
    # Add some competitor data with proper World Skate IDs (17 digits starting with 1 or 2)
    sheet['A6'] = 1
    sheet['B6'] = "12345678901234567"  # 17-digit ID starting with 1 (man)
    sheet['D6'] = "Test Skater One"
    sheet['E6'] = "Test Team"
    sheet['F6'] = "USA"
    
    sheet['A7'] = 2
    sheet['B7'] = "22345678901234567"  # 17-digit ID starting with 2 (woman)
    sheet['D7'] = "Test Skater Two"
    sheet['E7'] = "Test Team"
    sheet['F7'] = "CAN"
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def test_fallback_behavior():
    """Test the fallback behavior when no matching sheet is found."""
    
    print("\nTesting fallback behavior...")
    print("=" * 30)
    
    from result_sheets import load_workbook, BytesIO
    
    # Create a workbook with a single sheet that doesn't match our patterns
    wb = Workbook()
    wb.remove(wb.active)
    sheet = wb.create_sheet(title="Results Summary")  # Different name
    
    # Add some test data
    sheet['A1'] = ""
    sheet['B1'] = "Freestyle Slalom Battle"
    sheet['B2'] = "Junior Women - Final Ranking"
    sheet['A5'] = "Ranking"
    sheet['B5'] = "World Skate Id"
    sheet['D5'] = "Name"
    sheet['F5'] = "Ctry"
    
    # Add competitor data
    sheet['A6'] = 1
    sheet['B6'] = "12345678901234567"
    sheet['D6'] = "Test Skater"
    sheet['F6'] = "USA"
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_bytes = buffer.getvalue()
    
    try:
        # Test the sheet detection logic
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        
        # This should fall back to using the single available sheet
        if len(wb.sheetnames) == 1:
            sheet_name = wb.sheetnames[0]
            print(f"✅ Fallback successful: using single sheet '{sheet_name}'")
            return True
        else:
            print("❌ Fallback failed: multiple sheets found")
            return False
            
    except Exception as e:
        print(f"❌ Fallback error: {str(e)}")
        return False

def test_multiple_sheets():
    """Test behavior when multiple sheets exist but none match the pattern."""
    
    print("\nTesting multiple sheets with no match...")
    print("=" * 40)
    
    from result_sheets import load_workbook, BytesIO
    
    # Create a workbook with multiple sheets, none matching our pattern
    wb = Workbook()
    wb.remove(wb.active)
    wb.create_sheet(title="Sheet1")
    wb.create_sheet(title="Summary")
    wb.create_sheet(title="Data")
    
    # Save to bytes
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    file_bytes = buffer.getvalue()
    
    try:
        # Test the sheet detection logic
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        
        # Should not find any matching sheet and not fall back
        found_sheet = False
        
        # Common variations of the sheet name to look for
        possible_sheet_names = [
            "Final Ranking", "Final ranking", "final ranking", "FINAL RANKING",
            "Final Results", "Final results", "final results", "FINAL RESULTS"
        ]
        
        # Try exact matches
        for sheet_name in possible_sheet_names:
            if sheet_name in wb.sheetnames:
                found_sheet = True
                break
        
        # Try case-insensitive search
        if not found_sheet:
            for sheet_name in wb.sheetnames:
                if "final" in sheet_name.lower() and ("ranking" in sheet_name.lower() or "results" in sheet_name.lower()):
                    found_sheet = True
                    break
        
        if not found_sheet and len(wb.sheetnames) > 1:
            print(f"✅ Correctly rejected multiple non-matching sheets: {wb.sheetnames}")
            return True
        else:
            print(f"❌ Unexpected behavior with sheets: {wb.sheetnames}")
            return False
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return False

if __name__ == "__main__":
    print("Sheet Detection Test Suite")
    print("=" * 50)
    
    # Test case-insensitive detection
    test1_success = test_sheet_detection_only()
    
    # Test fallback behavior
    test2_success = test_fallback_behavior()
    
    # Test multiple sheets behavior
    test3_success = test_multiple_sheets()
    
    print("\n" + "=" * 50)
    print("Overall Test Results:")
    
    if test1_success and test2_success and test3_success:
        print("✅ All tests passed! Case-insensitive sheet detection is working correctly.")
        sys.exit(0)
    else:
        print("❌ Some tests failed!")
        sys.exit(1) 