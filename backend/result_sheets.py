import csv
import argparse
import os
import datetime
import base64
import time
import json
import requests
import msal
import unicodedata
import re
from registration_data import RegistrationData
from typing import List, Tuple, Optional, Set, Dict, Any
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

gs_dateOfBirthFormat = "%d/%m/%Y"

reg_data = RegistrationData("registration_responses.csv")

# Load secrets for OneDrive authentication
try:
    with open("secrets.json") as f:
        secrets = json.load(f)
except FileNotFoundError:
    print("Warning: secrets.json not found. OneDrive functionality will not be available.")
    secrets = {}

# Microsoft Graph settings
AUTHORITY = "https://login.microsoftonline.com/common"
SCOPES = ["Files.Read", "Files.Read.All"]
TOKEN_CACHE_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "cache", "token_cache.json"
)

# Authentication state
auth_state = {"flow": None, "token": None, "expires_at": 0, "is_authenticated": False}

def transliterate_text(text: str) -> str:
    """
    Transliterate text to remove accents and special characters.
    Based on the normalizeText function from the frontend.
    """
    if not text:
        return ""
    
    # Convert to string if not already
    text = str(text)
    
    # Replace accented characters with non-accented equivalents using NFD normalization
    # This handles most accented Latin characters
    normalized = unicodedata.normalize('NFD', text)
    # Remove combining characters (accents)
    normalized = re.sub(r'[\u0300-\u036f]', '', normalized)
    
    # Handle special characters that don't decompose properly with NFD
    # More efficient lookup table for non-standard transliterations
    transliterations = {
        # Special character pairs
        'æ': 'ae', 'œ': 'oe', 'ø': 'o', 'ö': 'o', 'ő': 'o',
        'å': 'a', 'ä': 'a', 'â': 'a', 'à': 'a', 'á': 'a',
        'ë': 'e', 'ê': 'e', 'è': 'e', 'é': 'e', 'ę': 'e',
        'ï': 'i', 'î': 'i', 'ì': 'i', 'í': 'i',
        'ü': 'u', 'û': 'u', 'ù': 'u', 'ú': 'u', 'ű': 'u',
        'ÿ': 'y', 'ý': 'y',
        'ß': 'ss', 'þ': 'th',
        
        # Slavic/Eastern European
        'č': 'c', 'ć': 'c', 'ċ': 'c',
        'š': 's', 'ś': 's', 'ŝ': 's',
        'ž': 'z', 'ź': 'z', 'ż': 'z',
        'ň': 'n', 'ń': 'n', 'ñ': 'n',
        'ř': 'r', 'ŕ': 'r', 'ŗ': 'r',
        'ď': 'd', 'đ': 'd', 'ð': 'd',
        'ť': 't', 'ț': 't', 'ŧ': 't',
        'ł': 'l', 'ŀ': 'l',
        
        # Additional characters that need special handling
        'Ł': 'L', 'Ø': 'O', 'Å': 'A', 'Æ': 'AE', 'Œ': 'OE',
        'Ö': 'O', 'Ü': 'U', 'Ä': 'A', 'Ñ': 'N', 'Ç': 'C',
        'Š': 'S', 'Ž': 'Z', 'Č': 'C', 'Ř': 'R', 'Ň': 'N',
        'Ť': 'T', 'Ď': 'D', 'Ľ': 'L', 'Ĺ': 'L', 'Ŕ': 'R',
        
        # Other special characters
        'ı': 'i', 'ȷ': 'j', 'ĸ': 'k',
        'ŉ': 'n', 'ſ': 's', 
        'ą': 'a'
    }
    
    # Apply transliterations
    result = ""
    for char in normalized:
        if char in transliterations:
            result += transliterations[char]
        else:
            result += char
    
    return result

def load_token_cache():
    """Load the token cache from base64-encoded JSON file if it exists."""
    print(f"Attempting to load token cache from: {TOKEN_CACHE_FILE}")
    if not os.path.exists(TOKEN_CACHE_FILE):
        print("Token cache file does not exist")
        return None
    try:
        encoded = open(TOKEN_CACHE_FILE, 'r').read()
        decoded = base64.b64decode(encoded).decode('utf-8')
        cache = msal.SerializableTokenCache()
        cache.deserialize(decoded)
        print("Successfully loaded token cache")
        return cache
    except Exception as e:
        print(f"Failed to load token cache: {e}")
        return None

def save_token_cache(cache):
    """Save the token cache to base64-encoded JSON file only if we have a valid token."""
    try:
        # Check if we have any accounts in the cache
        accounts = cache.find(msal.TokenCache.CredentialType.ACCOUNT)
        if not accounts:
            print("No accounts in cache, skipping save")
            return

        print(f"Attempting to save token cache to: {TOKEN_CACHE_FILE}")
        serialized = cache.serialize()
        encoded = base64.b64encode(serialized.encode('utf-8')).decode('utf-8')
        with open(TOKEN_CACHE_FILE, 'w') as f:
            f.write(encoded)
        print("Successfully saved token cache")
    except Exception as e:
        print(f"Failed to save token cache: {e}")

def get_user_token():
    """Get an access token using device code flow."""
    global auth_state
    now = time.time()
    if auth_state["token"] and now < auth_state["expires_at"]:
        print("Using cached token from memory")
        return auth_state["token"]

    print("No valid token in memory, attempting to load from cache")
    cache = load_token_cache() or msal.SerializableTokenCache()
    client = msal.PublicClientApplication(
        secrets["microsoft"]["client_id"],
        authority=AUTHORITY,
        token_cache=cache
    )

    accounts = client.get_accounts()
    if accounts:
        print(f"Found {len(accounts)} accounts in token cache")
        try:
            result = client.acquire_token_silent(SCOPES, account=accounts[0])
            if result and "access_token" in result:
                print("Successfully acquired token silently")
                auth_state.update({
                    "token": result["access_token"],
                    "expires_at": now + result.get("expires_in", 3600),
                    "is_authenticated": True
                })
                save_token_cache(cache)
                return result["access_token"]
            else:
                print("Silent token acquisition failed")
        except Exception as e:
            print(f"Error during silent token acquisition: {e}")

    raise Exception("Not authenticated. Please run the main application and authenticate first.")

def resolve_drive_item_ids(share_url: str) -> Tuple[str, str]:
    """Resolve and cache driveId and itemId from a 1drv.ms sharing URL."""
    token = get_user_token()
    b64 = base64.urlsafe_b64encode(share_url.encode('utf-8')).decode().rstrip('=')
    share_id = f"u!{b64}"
    meta_url = f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem?$select=id,parentReference"
    headers = {"Authorization": f"Bearer {token}"}
    print(f"Resolving drive and item IDs via: {meta_url}")
    resp = requests.get(meta_url, headers=headers)
    resp.raise_for_status()
    data = resp.json()
    drive_id = data["parentReference"]["driveId"]
    item_id = data["id"]
    print(f"Resolved driveId={drive_id}, itemId={item_id}")
    return drive_id, item_id

def list_onedrive_folder_contents(share_url: str) -> List[Dict[str, Any]]:
    """List the contents of a OneDrive folder from a sharing URL."""
    token = get_user_token()
    drive_id, item_id = resolve_drive_item_ids(share_url)
    
    # List children of the folder
    list_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/children"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json"
    }
    
    print(f"Listing folder contents from: {list_url}")
    resp = requests.get(list_url, headers=headers)
    resp.raise_for_status()
    data = resp.json()
    
    files = []
    for item in data.get("value", []):
        if item.get("file"):  # Only include files, not folders
            files.append({
                "name": item["name"],
                "id": item["id"],
                "size": item.get("size", 0),
                "lastModified": item.get("lastModifiedDateTime"),
                "downloadUrl": item.get("@microsoft.graph.downloadUrl"),
                "driveId": drive_id
            })
    
    print(f"Found {len(files)} files in folder")
    return files

def download_excel_file(drive_id: str, item_id: str) -> bytes:
    """Download an Excel file from OneDrive using drive and item IDs."""
    token = get_user_token()
    
    # Download content
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Cache-Control": "no-cache"
    }
    
    print(f"Downloading Excel file from: {url}")
    resp = requests.get(url, headers=headers, stream=True)
    resp.raise_for_status()
    
    # Read content
    content = bytearray()
    for chunk in resp.iter_content(chunk_size=1024*1024):
        if chunk:
            content.extend(chunk)
    
    if len(content) == 0:
        raise Exception("Received empty file from OneDrive")
    
    print(f"Downloaded {len(content)} bytes")
    return bytes(content)

def process_excel_final_ranking(file_bytes: bytes, filename: str) -> Tuple[str, str, List[List[str]]]:
    """Process an Excel file's 'Final Ranking' sheet and extract results."""
    extracted_rows: List[List[str]] = []
    discipline: str = ""
    category: str = ""
    
    try:
        # Load the workbook
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        
        # Find the "Final Ranking" sheet (case-insensitive)
        final_ranking_sheet = None
        final_ranking_sheet_name = None
        
        # Common variations of the sheet name to look for
        possible_sheet_names = [
            "Final Ranking",
            "Final ranking", 
            "final ranking",
            "FINAL RANKING",
            "Final Results",
            "Final results",
            "final results",
            "FINAL RESULTS"
        ]
        
        # First, try exact matches with the common variations
        for sheet_name in possible_sheet_names:
            if sheet_name in wb.sheetnames:
                final_ranking_sheet = wb[sheet_name]
                final_ranking_sheet_name = sheet_name
                print(f"Found sheet: '{sheet_name}' in {filename}")
                break
        
        # If no exact match found, try case-insensitive search
        if final_ranking_sheet is None:
            for sheet_name in wb.sheetnames:
                if "final" in sheet_name.lower() and ("ranking" in sheet_name.lower() or "results" in sheet_name.lower()):
                    final_ranking_sheet = wb[sheet_name]
                    final_ranking_sheet_name = sheet_name
                    print(f"Found sheet (case-insensitive): '{sheet_name}' in {filename}")
                    break
        
        # If still no match, check if there's only one sheet and use it
        if final_ranking_sheet is None:
            if len(wb.sheetnames) == 1:
                final_ranking_sheet = wb[wb.sheetnames[0]]
                final_ranking_sheet_name = wb.sheetnames[0]
                print(f"Using only available sheet: '{final_ranking_sheet_name}' in {filename}")
            else:
                print(f"Warning: No 'Final Ranking' sheet found in {filename}. Available sheets: {wb.sheetnames}")
                return "", "", []
        
        sheet = final_ranking_sheet
        
        # Convert sheet to DataFrame for easier processing
        # First, let's read the sheet data into a list of lists
        sheet_data = []
        for row in sheet.iter_rows(values_only=True):
            sheet_data.append(list(row) if row else [])
        
        # Find discipline and category from the first few rows
        discipline_found = False
        for i in range(min(10, len(sheet_data))):
            row = sheet_data[i]
            if not row:
                continue
                
            for j in range(min(6, len(row))):
                cell_value = row[j]
                if isinstance(cell_value, str) and cell_value.strip():
                    cell_text = cell_value.strip()
                    
                    # Check for discipline keywords
                    discipline_keywords = ["Classic", "Battle", "Freejump", "Freestyle Slalom"]
                    for keyword in discipline_keywords:
                        if keyword in cell_text:
                            discipline = cell_text
                            discipline_found = True
                            
                            # Look for category in nearby cells
                            # Check the next row in the same column
                            if i + 1 < len(sheet_data) and j < len(sheet_data[i + 1]):
                                next_cell = sheet_data[i + 1][j]
                                if isinstance(next_cell, str) and next_cell.strip():
                                    category = next_cell.strip()
                            break
                    
                    if discipline_found:
                        break
            
            if discipline_found:
                break
        
        # If no discipline found from content, try to extract from filename
        if not discipline:
            filename_lower = filename.lower()
            if "classic" in filename_lower:
                discipline = "Classic"
            elif "battle" in filename_lower:
                discipline = "Battle"
            elif "freejump" in filename_lower:
                discipline = "Freejump"
            else:
                discipline = filename.split('.')[0]  # Use filename without extension
        
        # Find the header row with ranking data
        header_found = False
        data_start_row = 0
        rank_idx = None
        id_idx = None
        name_idx = None
        ctry_idx = None
        
        for i, row in enumerate(sheet_data):
            if not row:
                continue
                
            # Convert row to lowercase for case-insensitive comparison
            row_lower = [str(cell).lower() if cell is not None else "" for cell in row]
            
            # Look for header indicators
            for j, cell in enumerate(row_lower):
                if cell in ['rank', 'ranking', 'place']:
                    rank_idx = j
                elif 'id' in cell and 'skate' in cell:
                    id_idx = j
                elif cell == 'name':
                    name_idx = j
                elif cell in ['ctry', 'country', 'nationality']:
                    ctry_idx = j
            
            # Check if we found the essential columns
            if rank_idx is not None and name_idx is not None and ctry_idx is not None:
                header_found = True
                data_start_row = i + 1
                break
        
        if not header_found:
            print(f"Warning: Could not find header row in {filename} (sheet: '{final_ranking_sheet_name}')")
            return discipline, category, []
        
        # Process data rows
        data_started = False
        for i in range(data_start_row, len(sheet_data)):
            row = sheet_data[i]
            if not row or len(row) <= max(rank_idx or 0, name_idx or 0, ctry_idx or 0):
                continue
            
            # Get rank
            rank_cell = row[rank_idx] if rank_idx is not None else None
            if rank_cell is None:
                continue
                
            rank_str = str(rank_cell).strip()
            
            # Skip non-numeric ranks unless we've started processing data
            if not rank_str.isdigit():
                if data_started:
                    break  # End of data
                else:
                    continue  # Haven't started yet
            
            rank = int(rank_str)
            if rank == 1:
                data_started = True
            
            # Get other required fields
            full_name = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] is not None else ""
            ctry = str(row[ctry_idx]).strip() if ctry_idx is not None and row[ctry_idx] is not None else ""
            
            if not full_name:
                continue
            
            # Parse name
            family_name, first_name = parse_name(full_name.split(), ctry)
            
            # Get World Skate ID (using original names before transliteration)
            if id_idx is not None and row[id_idx] is not None:
                skater_id = str(row[id_idx]).strip()
            else:
                # Try to find ID in registration data using original names
                skater_data = reg_data.get_all_data_by_name(family_name, first_name)
                if skater_data:
                    skater_id = skater_data['id']
                else:
                    skater_id = "N/A"
                    print(f"Skater {full_name} ({ctry}) not found in registration list")
            
            # Get birthdate (using original names before transliteration)
            if len(skater_id) < 5:
                print(f"Skater {full_name} does not have a proper World Skate ID, trying to find birthdate in the registration list")
                skater_data = reg_data.get_all_data_by_name(family_name, first_name)
                if skater_data:
                    birthdate = skater_data['date_of_birth'].strftime(gs_dateOfBirthFormat)
                    skater_id = skater_data['id']
                else:
                    raise ValueError(f"Skater {full_name} not found in registration list")
            else:
                skater_data = reg_data.get_by_id(skater_id)
                if skater_data:
                    birthdateDt = skater_data['date_of_birth']
                    if birthdateDt is not None:
                        birthdate = birthdateDt.strftime(gs_dateOfBirthFormat)
                    else:
                        birthdate = "N/A"
                        print(f"Skater {skater_id} {full_name} does not have a birthdate in the registration list")
                else:
                    raise ValueError(f"Skater {full_name} not found in registration list")
            
            # Apply transliteration to remove accents and special characters (for output only)
            first_name_transliterated = transliterate_text(first_name)
            family_name_transliterated = transliterate_text(family_name)
            
            gender = skater_gender_from_id(skater_id)
            
            extracted_rows.append([
                rank_str,
                skater_id,
                first_name_transliterated,
                family_name_transliterated,
                gender,
                birthdate,
                ctry
            ])
        
        print(f"Processed {len(extracted_rows)} competitors from {filename} (sheet: '{final_ranking_sheet_name}')")
        return discipline, category, extracted_rows
        
    except Exception as e:
        print(f"Error processing Excel file {filename}: {str(e)}")
        return "", "", []

# Placeholder birthdate function (replace with real implementation)
def wssid_to_birthdate(skater_id: str) -> datetime.date:
    return reg_data.get_date_of_birth(skater_id, gs_dateOfBirthFormat)

# Parsing skater name based on nationality
def parse_name(name_parts: List[str], ctry: str) -> Tuple[str, str]:
    spanish_speaking: Set[str] = {'ESP', 'ARG', 'MEX', 'COL', 'CHI', 'PER', 'VEN', 'URU', 'CUB'}
    east_asian: Set[str] = {'CHN', 'KOR'}

    if len(name_parts) == 2:
        family_name, first_name = name_parts
    elif ctry.upper() in spanish_speaking:
        first_name = name_parts[-1]
        family_name = ' '.join(name_parts[:-1])
    elif ctry.upper() in {'CHN', 'KOR'}:
        family_name = name_parts[0]
        first_name = ' '.join(name_parts[1:])
    else:
        family_name = name_parts[0]
        first_name = ' '.join(name_parts[1:])

    return family_name, first_name 

# Extracting gender from ID
def skater_gender_from_id(skater_id: str) -> str:
    return "man" if skater_id.startswith('1') else "woman" if skater_id.startswith('2') else ""

def process_csv(filepath: str) -> Tuple[str, str, List[List[str]]]:
    """Original CSV processing function - kept for backward compatibility."""
    extracted_rows: List[List[str]] = []
    discipline: str = ""
    category: str = ""

    with open(filepath, newline='', encoding='cp1252') as csvfile:
        reader = csv.reader(csvfile)

        header_found: bool = False
        discipline_found: bool = False
        data_started: bool = False
        rank_idx: Optional[int] = None
        id_idx: Optional[int] = None
        name_idx: Optional[int] = None
        ctry_idx: Optional[int] = None

        for row in reader:
            if row and row[0] in {"Picture", "Results", "Detailed results"}:
                continue

            if not discipline_found and row:
                # Check both first and second columns for discipline
                discipline_keywords = ["Classic", "Battle", "Freejump"]
                for col in [0, 1]:
                    if col < len(row) and any(keyword in row[col] for keyword in discipline_keywords):
                        discipline = row[col].strip()
                        # Get category from the next row in the same column
                        next_row = next(reader, [])
                        category = next_row[col].strip() if col < len(next_row) else ""
                        discipline_found = True
                        break
                if discipline_found:
                    continue

            if not header_found and row:
                # Convert row to lowercase for case-insensitive comparison
                row_lower: List[str] = [col.lower() for col in row]
                
                # Find indices for each required column with flexible naming
                rank_idx = None
                id_idx = None
                name_idx = None
                ctry_idx = None
                
                for i, col in enumerate(row_lower):
                    if col in ['rank', 'ranking', 'place']:
                        rank_idx = i
                    elif 'id' in col:
                        id_idx = i
                    elif col == 'name':
                        name_idx = i
                    elif col in ['ctry', 'country', 'nationality']:
                        ctry_idx = i
                
                # Check if we found all required columns
                if all(idx is not None for idx in [rank_idx, id_idx, name_idx, ctry_idx]):
                    header_found = True
                    continue
                # If the ID is missing, we can still try to find it in the registration list using the name and surname
                if id_idx is None and name_idx is not None and ctry_idx is not None:
                    header_found = True
                    continue

            if header_found:
                if len(row) <= rank_idx:
                    continue

                rank: str = row[rank_idx].strip()

                # If the rank is not a digit, we need to skip the row
                if not rank.isdigit():
                    if data_started:
                        # Data ended
                        break
                    else:
                        # Data not started yet
                        continue

                if rank == "1":
                    data_started = True
                full_name: str = row[name_idx].strip()
                ctry: str = row[ctry_idx].strip()
                family_name, first_name = parse_name(full_name.split(), ctry)
                
                if id_idx is not None:
                    skater_id: str = row[id_idx].strip()
                else:
                    # If the ID is missing, we need to find it in the registration list using the name and surname (original names)
                    skater_data = reg_data.get_all_data_by_name(family_name, first_name)
                    if skater_data:
                        skater_id = skater_data['id']
                    else:
                        skater_id = "N/A"
                        print(f"Skater {full_name} ({ctry}) not found in registration list")

                if len(skater_id) < 5:
                    print(f"Skater {full_name} does not have a proper World Skate ID, trying to find birthdate in the registration list")
                    # If the skater does not have a proper World Skate ID, we need to find their birthdate
                    # from the registration list using their name and surname (original names)
                    skater_data = reg_data.get_all_data_by_name(family_name, first_name)
                    if skater_data:
                        birthdate = skater_data['date_of_birth'].strftime(gs_dateOfBirthFormat)
                        skater_id = skater_data['id']
                    else:
                        # Could not find the skater in the registration list, this is an error
                        raise ValueError(f"Skater {full_name} not found in registration list")
                else:
                    #print(f"Skater ID: {skater_id} {full_name}")
                    skater_data = reg_data.get_by_id(skater_id)
                    if skater_data:
                        birthdateDt = skater_data['date_of_birth']
                        if birthdateDt is not None:
                            birthdate = birthdateDt.strftime(gs_dateOfBirthFormat)
                        else:
                            birthdate = "N/A"
                            print(f"Skater {skater_id} {full_name} does not have a birthdate in the registration list")
                    else:
                        # Could not find the skater in the registration list, this is an error
                        raise ValueError(f"Skater {full_name} not found in registration list")
                
                # Apply transliteration to remove accents and special characters (for output only)
                first_name_transliterated = transliterate_text(first_name)
                family_name_transliterated = transliterate_text(family_name)
                
                gender: str = skater_gender_from_id(skater_id)

                extracted_rows.append([
                    rank,
                    skater_id,
                    first_name_transliterated,
                    family_name_transliterated,
                    gender,
                    birthdate,
                    ctry
                ])

    return discipline, category, extracted_rows

def process_onedrive_folder(share_url: str) -> List[Tuple[str, str, List[List[str]]]]:
    """Process all Excel files in a OneDrive folder."""
    print(f"Processing OneDrive folder: {share_url}")
    
    # List folder contents
    files = list_onedrive_folder_contents(share_url)
    
    # Filter for Excel files
    excel_files = [f for f in files if f["name"].lower().endswith(('.xlsx', '.xls'))]
    print(f"Found {len(excel_files)} Excel files")
    
    results = []
    
    for file_info in excel_files:
        print(f"Processing {file_info['name']}")
        try:
            # Download the file
            file_bytes = download_excel_file(file_info["driveId"], file_info["id"])
            
            # Process the Excel file
            discipline, category, rows = process_excel_final_ranking(file_bytes, file_info["name"])
            
            if rows:  # Only add if we got results
                results.append((discipline, category, rows))
            else:
                print(f"No results found in {file_info['name']}")
                
        except Exception as e:
            print(f"Error processing {file_info['name']}: {str(e)}")
            continue
    
    return results

def main() -> None:
    parser: argparse.ArgumentParser = argparse.ArgumentParser(description="Process skating competition files from OneDrive or local CSV files.")
    parser.add_argument("source", help="OneDrive folder URL (1drv.ms link) or path to CSV files/directory")
    parser.add_argument("--output", "-o", default="combined_results.csv", help="Output CSV file name")
    args: argparse.Namespace = parser.parse_args()

    output_file: str = args.output
    results = []
    
    # Check if source is a OneDrive URL
    if "1drv.ms" in args.source:
        print("Processing OneDrive folder...")
        try:
            results = process_onedrive_folder(args.source)
        except Exception as e:
            print(f"Error processing OneDrive folder: {str(e)}")
            print("Make sure you are authenticated. Run the main application first to authenticate.")
            return
    else:
        # Handle local CSV files (original functionality)
        print("Processing local CSV files...")
        csv_files = []
        
        if os.path.isdir(args.source):
            csv_files = [os.path.join(args.source, f) for f in os.listdir(args.source) if f.endswith('.csv')]
        elif os.path.isfile(args.source):
            csv_files = [args.source]
        else:
            # Assume it's a list of files
            csv_files = [args.source]
        
        for file in csv_files:
            if os.path.exists(file):
                print(f"Processing {file}")
                try:
                    discipline, category, rows = process_csv(file)
                    if not discipline:
                        discipline = os.path.basename(file).split(".")[0]
                    results.append((discipline, category, rows))
                except Exception as e:
                    print(f"Error processing {file}: {str(e)}")
            else:
                print(f"File not found: {file}")
    
    if not results:
        print("No results to write.")
        return
    
    # Write combined results to CSV
    with open(output_file, 'w', newline='', encoding='utf-8') as csv_out:
        writer: csv.writer = csv.writer(csv_out)
        
        # Process each result and write its contents
        for i, (discipline, category, rows) in enumerate(results):
            print(f"Writing results for {discipline} - {category}")
            
            # Write discipline and category
            writer.writerow(["DISCIPLINE", discipline])
            writer.writerow(["CATEGORY", category])
            
            # Write column headers for this discipline
            writer.writerow(["RANK", "WORLD SKATE SKATER ID", "FIRST NAME", "FAMILY NAME", "GENDER", "BIRTHDATE DD/MM/YYYY", "NATIONALITY"])
            
            # Write the data rows
            writer.writerows(rows)
            
            # Add delimiter lines between disciplines (except after the last one)
            if i < len(results) - 1:
                writer.writerow([])  # Empty line
                writer.writerow([])  # Empty line

    print(f"Combined results saved to '{output_file}'")
    print(f"Processed {len(results)} disciplines with {sum(len(rows) for _, _, rows in results)} total competitors")

if __name__ == "__main__":
    main()
